#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将「人情笔记_*.xlsx」历史数据迁移进 Gift Ledger 的 SQLite 数据库。

特点：
  - 直接写 SQLite 文件，绕开应用 Web 导入（页面导入也能用，但本脚本迁移更完整可靠）
  - 自动为缺失的「亲友类型 / 事由」创建记录，避免关联信息丢失
  - 每个用户：先补默认 reasons / contact_types，再对 Excel 里出现的新类型/事由自动创建
  - 礼簿、亲友自动创建；已存在且完全一致的行跳过（幂等，可重复执行）
  - 日期统一按 YYYY-MM-DD 处理，绝不触发时区换算

注意：应用 importHelper.formatDate 曾有时区 bug（礼簿日期/随礼日期带 "00:00:43"，
在 +8 时区下会被 toISOString 回退一天），已在 e4ccdf4 修复。本脚本因直接解析、
不走 Web 导入，从一开始就绕开了该问题；同时它在类型/事由补全上比 Web 导入更完整，
因此仍推荐用它来迁移历史 Excel。

用法：
  python migrate_excel_to_sqlite.py \
      --db /vol1/1000/docker/gift-ledger/data/gift_ledger.db \
      --excel 人情笔记_示例_20260716_2200.xlsx \
      --user your_username

依赖：
  pip install openpyxl
（仅用标准库 sqlite3，无需 better-sqlite3）

前提：目标用户须先在应用内注册（脚本按用户名解析 user_id，不处理密码）。
"""
import argparse
import re
import sqlite3
import sys
from datetime import datetime, timedelta

DEFAULT_REASONS = ['丧事', '孝敬', '其它', '压岁', '生日', '生子', '婚礼']
DEFAULT_TYPES = ['领导', '其它', '同学', '朋友', '亲戚', '同事']


def parse_date(value):
    """返回 YYYY-MM-DD 或 None。不做任何时区换算。"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        d = datetime(1899, 12, 30) + timedelta(days=int(value))
        return d.strftime('%Y-%m-%d')
    s = str(value).strip()
    m = re.match(r'^(\d{4}-\d{2}-\d{2})', s)  # 兼容 "2021-05-03 00:00:43"
    if m:
        return m.group(1)
    try:
        return datetime.strptime(s, '%Y/%m/%d').strftime('%Y-%m-%d')
    except Exception:
        return None


def parse_datetime(value):
    """返回 YYYY-MM-DD HH:MM:SS 或 None。"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        d = datetime(1899, 12, 30) + timedelta(days=value)
        return d.strftime('%Y-%m-%d %H:%M:%S')
    s = str(value).strip()
    if re.match(r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$', s):
        return s
    if re.match(r'^\d{4}-\d{2}-\d{2}$', s):
        return s + ' 00:00:00'
    return None


def get_or_create(cur, user_id, table, name):
    if not name:
        return None
    name = str(name).strip()
    cur.execute(f"SELECT id FROM {table} WHERE user_id=? AND name=?", (user_id, name))
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute(f"INSERT INTO {table} (user_id, name) VALUES (?,?)", (user_id, name))
    return cur.lastrowid


def ensure_defaults(cur, user_id):
    for n in DEFAULT_REASONS:
        get_or_create(cur, user_id, 'reasons', n)
    for n in DEFAULT_TYPES:
        get_or_create(cur, user_id, 'contact_types', n)


def get_or_create_book(cur, user_id, name, date):
    cur.execute("SELECT id FROM gift_books WHERE user_id=? AND name=?", (user_id, name))
    row = cur.fetchone()
    if row:
        return row[0]
    cur.execute("INSERT INTO gift_books (user_id, name, date) VALUES (?,?,?)", (user_id, name, date))
    return cur.lastrowid


def ensure_contact(cur, user_id, name, type_id):
    if not name:
        return
    cur.execute("SELECT id, type_id FROM contacts WHERE user_id=? AND name=?", (user_id, name))
    row = cur.fetchone()
    if row is None:
        cur.execute("INSERT INTO contacts (user_id, name, type_id) VALUES (?,?,?)", (user_id, name, type_id))
    elif type_id and not row[1]:
        cur.execute("UPDATE contacts SET type_id=? WHERE id=?", (type_id, row[0]))


def import_received(cur, user_id, rows):
    succ = fail = 0
    errs = []
    for i, r in enumerate(rows, start=2):
        try:
            contact = r.get('亲友姓名')
            ctype = r.get('亲友类型')
            amount = r.get('金额')
            book = r.get('所属礼簿')
            bdate = parse_date(r.get('礼簿日期'))
            notes = r.get('备注')
            created = parse_datetime(r.get('创建时间'))
            notes = str(notes).strip() if notes is not None else None
            notes = notes or None
            if not contact or amount is None or not book or not bdate:
                fail += 1
                errs.append(f"第{i}行：缺少必填字段")
                continue
            try:
                amount = float(amount)
            except Exception:
                fail += 1
                errs.append(f"第{i}行：金额无效")
                continue
            if amount <= 0:
                fail += 1
                errs.append(f"第{i}行：金额<=0")
                continue
            type_id = get_or_create(cur, user_id, 'contact_types', ctype) if ctype else None
            book_id = get_or_create_book(cur, user_id, str(book).strip(), bdate)
            # 幂等去重
            cur.execute(
                "SELECT id FROM gifts_received WHERE user_id=? AND contact_name=? AND amount=? "
                "AND gift_book_id=? AND gift_book_date=? AND (notes IS ? OR notes=?) AND (created_at IS ? OR created_at=?)",
                (user_id, contact, amount, book_id, bdate, notes, notes, created, created))
            if cur.fetchone():
                fail += 1
                errs.append(f"第{i}行：{contact} {amount}元 {bdate} 重复，跳过")
                continue
            cur.execute(
                "INSERT INTO gifts_received (user_id, contact_name, contact_type_id, amount, gift_book_id, gift_book_date, notes, created_at) "
                "VALUES (?,?,?,?,?,?,?,?)",
                (user_id, contact, type_id, amount, book_id, bdate, notes, created))
            succ += 1
            ensure_contact(cur, user_id, contact, type_id)
        except Exception as e:
            fail += 1
            errs.append(f"第{i}行：{e}")
    return succ, fail, errs


def import_given(cur, user_id, rows):
    succ = fail = 0
    errs = []
    for i, r in enumerate(rows, start=2):
        try:
            contact = r.get('亲友姓名')
            ctype = r.get('亲友类型')
            amount = r.get('金额')
            reason = r.get('事由')
            gdate = parse_date(r.get('随礼日期'))
            notes = r.get('备注')
            created = parse_datetime(r.get('创建时间'))
            notes = str(notes).strip() if notes is not None else None
            notes = notes or None
            if not contact or amount is None or not gdate:
                fail += 1
                errs.append(f"第{i}行：缺少必填字段")
                continue
            try:
                amount = float(amount)
            except Exception:
                fail += 1
                errs.append(f"第{i}行：金额无效")
                continue
            if amount <= 0:
                fail += 1
                errs.append(f"第{i}行：金额<=0")
                continue
            reason_id = get_or_create(cur, user_id, 'reasons', reason) if reason else None
            type_id = get_or_create(cur, user_id, 'contact_types', ctype) if ctype else None
            cur.execute(
                "SELECT id FROM gifts_given WHERE user_id=? AND contact_name=? AND amount=? "
                "AND gift_date=? AND (reason_id IS ? OR reason_id=?) AND (notes IS ? OR notes=?) AND (created_at IS ? OR created_at=?)",
                (user_id, contact, amount, gdate, reason_id, reason_id, notes, notes, created, created))
            if cur.fetchone():
                fail += 1
                errs.append(f"第{i}行：{contact} {amount}元 {gdate} 重复，跳过")
                continue
            cur.execute(
                "INSERT INTO gifts_given (user_id, contact_name, contact_type_id, amount, reason_id, gift_date, notes, created_at) "
                "VALUES (?,?,?,?,?,?,?,?)",
                (user_id, contact, type_id, amount, reason_id, gdate, notes, created))
            succ += 1
            ensure_contact(cur, user_id, contact, type_id)
        except Exception as e:
            fail += 1
            errs.append(f"第{i}行：{e}")
    return succ, fail, errs


def load_sheet_rows(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(h).strip() if h is not None else '' for h in rows[0]]
    return [dict(zip(header, [c for c in r])) for r in rows[1:] if any(c is not None for c in r)]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--db', required=True, help='SQLite 数据库文件路径')
    ap.add_argument('--excel', required=True, help='要导入的 Excel 文件')
    ap.add_argument('--user', required=True, help='目标用户名（须已在应用内注册）')
    args = ap.parse_args()

    from openpyxl import load_workbook
    conn = sqlite3.connect(args.db)
    cur = conn.cursor()
    cur.execute("SELECT id FROM users WHERE username=?", (args.user,))
    row = cur.fetchone()
    if not row:
        conn.close()
        print(f"[错误] 用户 '{args.user}' 不存在，请先在应用内注册该用户。", file=sys.stderr)
        sys.exit(2)
    user_id = row[0]
    print(f"[用户] {args.user} -> user_id={user_id}")

    ensure_defaults(cur, user_id)
    conn.commit()

    wb = load_workbook(args.excel, read_only=True, data_only=True)
    received_sheet = wb['收礼'] if '收礼' in wb.sheetnames else None
    given_sheet = wb['随礼'] if '随礼' in wb.sheetnames else None

    if received_sheet:
        s, f, e = import_received(cur, user_id, load_sheet_rows(received_sheet))
        conn.commit()
        print(f"[收礼] 成功 {s} 行，跳过/失败 {f} 行")
        for msg in e[:20]:
            print("   -", msg)
    else:
        print("[收礼] 未找到「收礼」sheet，跳过")

    if given_sheet:
        s, f, e = import_given(cur, user_id, load_sheet_rows(given_sheet))
        conn.commit()
        print(f"[随礼] 成功 {s} 行，跳过/失败 {f} 行")
        for msg in e[:20]:
            print("   -", msg)
    else:
        print("[随礼] 未找到「随礼」sheet，跳过")

    conn.close()
    print("[完成] 迁移结束。")


if __name__ == '__main__':
    main()
